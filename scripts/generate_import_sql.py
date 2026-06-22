#!/usr/bin/env python3
"""
Generate an applications INSERT SQL file from Lauren's scholarship tracker spreadsheet.

Usage:
    python3 scripts/generate_import_sql.py [--user-id 1] [--output docs/import.sql]

Defaults to user_id=1 (dev). Pass the real production user_profiles.id for prod imports.

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

SPREADSHEET = Path(__file__).parent.parent / "docs" / "LaurenScholarshipOrganizerTracker.xlsx"
DEFAULT_OUTPUT = Path(__file__).parent.parent / "docs" / "applications_import.sql"
PLACEHOLDER_DATE = "2099-12-31"

STATUS_MAP = {
    "submitted": "Submitted",
    "in progress": "In Progress",
    "not started": "Not Started",
    "no started": "Not Started",
    "awarded": "Awarded",
    "not awarded": "Not Awarded",
}

LEVEL_TO_TARGET = {
    "Y": "Need",
}


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")[:80]


def esc(val) -> str:
    if val is None:
        return "NULL"
    cleaned = str(val).replace("'", "''").replace("\n", " ").replace("\r", "")
    return f"'{cleaned}'"


def fmt_date(val) -> str:
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    return "NULL"


def fmt_money(val) -> str:
    if val is None:
        return "NULL"
    try:
        f = float(val)
        return "NULL" if f == 0 else str(f)
    except (ValueError, TypeError):
        return "NULL"


def map_status(raw) -> str:
    if raw is None:
        return "Not Started"
    return STATUS_MAP.get(str(raw).strip().lower(), "Not Started")


def parse_tab1(ws) -> list[dict]:
    """Rows 5-7 of the TEMPLATE long scholarship list tab."""
    rows = []
    clean_names = {}

    for row_idx in range(5, 8):
        row = ws[row_idx]
        cells = {cell.column_letter: cell for cell in row}

        a = cells.get("A")
        if not a or not a.value:
            continue

        name = str(a.value).strip()
        app_link = None
        if cells.get("H") and cells["H"].hyperlink:
            app_link = cells["H"].hyperlink.target
        elif cells.get("O") and cells["O"].hyperlink:
            app_link = cells["O"].hyperlink.target
        elif a.hyperlink:
            app_link = a.hyperlink.target

        deadline = cells.get("C") and cells["C"].value
        open_date = cells.get("I") and cells["I"].value
        amount = cells.get("F") and cells["F"].value
        theme = cells.get("K") and cells["K"].value
        status_raw = cells.get("B") and cells["B"].value

        rows.append({
            "scholarship_name": name,
            "application_link": app_link,
            "due_date": deadline if isinstance(deadline, datetime) else None,
            "open_date": open_date if isinstance(open_date, datetime) else None,
            "min_award": amount,
            "max_award": amount,
            "theme": str(theme).strip() if theme else None,
            "status": map_status(status_raw),
            "target_type": "Merit",
        })

    return rows


def parse_tab2(ws) -> list[dict]:
    """All scholarship rows from the College Only list tab."""
    clean_names = {
        "Check this out Phi Theta Kappa. Has fee to join": "Phi Theta Kappa",
        "https://www.gemfellowship.org/gem-fellowship-program/": "GEM Fellowship Program",
        "https://www.afcea.org/stem-majors-scholarships": "AFCEA STEM Majors Scholarships",
    }

    rows = []
    for row in ws.iter_rows(min_row=2):
        cells = {cell.column_letter: cell for cell in row}
        a = cells.get("A")
        if not a or not a.value:
            continue

        raw_name = str(a.value).strip()
        name = clean_names.get(raw_name, raw_name)

        # Application link: prefer hyperlink on col A, then text in col F if it's a URL
        app_link = a.hyperlink.target if a.hyperlink else None
        f_cell = cells.get("F")
        if not app_link and f_cell and isinstance(f_cell.value, str) and f_cell.value.startswith("http"):
            app_link = f_cell.value

        deadline_cell = cells.get("D")
        deadline = deadline_cell.value if deadline_cell else None

        open_date_cell = cells.get("E")
        open_date = open_date_cell.value if open_date_cell else None

        # Amount: col I normally; Lab Roots row has amount in col F
        i_cell = cells.get("I")
        amount_val = i_cell.value if i_cell else None
        if amount_val is None and f_cell and isinstance(f_cell.value, (int, float)):
            amount_val = f_cell.value
        if isinstance(amount_val, str):
            amount_val = None

        eligibility_cell = cells.get("L")
        requirements = str(eligibility_cell.value).strip() if eligibility_cell and eligibility_cell.value else None
        if raw_name == "Check this out Phi Theta Kappa. Has fee to join":
            requirements = (requirements + " Has a membership fee to join.").strip() if requirements else "Has a membership fee to join."

        need_cell = cells.get("M")
        need_val = str(need_cell.value).strip().upper() if need_cell and need_cell.value else None
        target_type = "Need" if need_val == "Y" else "Merit"

        status_cell = cells.get("B")
        status_raw = status_cell.value if status_cell else None

        rows.append({
            "scholarship_name": name,
            "application_link": app_link,
            "due_date": deadline if isinstance(deadline, datetime) else None,
            "open_date": open_date if isinstance(open_date, datetime) else None,
            "min_award": amount_val,
            "max_award": amount_val,
            "requirements": requirements,
            "status": map_status(status_raw),
            "target_type": target_type,
        })

    return rows


def build_sql(rows: list[dict], user_id: int) -> str:
    lines = [
        "-- applications import generated from Lauren's scholarship tracker",
        f"-- Generated: {datetime.now().strftime('%Y-%m-%d')}",
        f"-- user_id: {user_id}",
        "",
        "BEGIN;",
        "",
    ]

    cols = (
        "user_id, scholarship_name, application_link, "
        "due_date, open_date, min_award, max_award, "
        "theme, requirements, status, target_type"
    )

    for r in rows:
        due = fmt_date(r.get("due_date")) if r.get("due_date") else f"'{PLACEHOLDER_DATE}'"
        vals = ", ".join([
            str(user_id),
            esc(r["scholarship_name"]),
            esc(r.get("application_link")),
            due,
            fmt_date(r.get("open_date")),
            fmt_money(r.get("min_award")),
            fmt_money(r.get("max_award")),
            esc(r.get("theme")),
            esc(r.get("requirements")),
            esc(r["status"]),
            esc(r["target_type"]),
        ])
        lines.append(f"-- {r['scholarship_name']}")
        lines.append(f"INSERT INTO applications ({cols})")
        lines.append(f"VALUES ({vals});")
        lines.append("")

    lines.append("COMMIT;")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Generate applications import SQL from scholarship spreadsheet.")
    parser.add_argument("--user-id", type=int, default=1, help="user_profiles.id to assign applications to (default: 1)")
    parser.add_argument("--spreadsheet", type=Path, default=SPREADSHEET, help="Path to the .xlsx file")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output .sql file path")
    args = parser.parse_args()

    if not args.spreadsheet.exists():
        sys.exit(f"Spreadsheet not found: {args.spreadsheet}\nPass --spreadsheet <path>")

    wb = openpyxl.load_workbook(args.spreadsheet, data_only=True)
    tab1_rows = parse_tab1(wb["TEMPLATE long scholarship list "])
    tab2_rows = parse_tab2(wb["College Only list "])
    all_rows = tab1_rows + tab2_rows

    sql = build_sql(all_rows, args.user_id)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql)

    print(f"Wrote {len(all_rows)} applications → {args.output}")
    for r in all_rows:
        print(f"  {r['status']:12s}  {r['scholarship_name']}")


if __name__ == "__main__":
    main()
